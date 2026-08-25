"""
update.py
---------
'데이터 RAW' 폴더 내 모든 .xlsx 파일을 읽어 dashboard.html 데이터를 자동 갱신합니다.
- 새 파일을 '데이터 RAW' 폴더에 추가한 뒤 이 스크립트를 실행하면 됩니다
"""

import openpyxl, glob, re, json, calendar
from collections import defaultdict
from datetime import datetime, timedelta

DASHBOARD = 'dashboard.html'
DATA_DIR  = '데이터 RAW'
INDEX     = 'index.html'

# ── 유틸 ────────────────────────────────────────────────────────

def n(v):
    try: return float(str(v).replace(',', ''))
    except: return 0

def to_date(v):
    if v is None: return ''
    if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
    s = str(v)
    return s[:10] if len(s) >= 10 else ''

def shorten(name, maxlen=22):
    return name[:maxlen].strip()

def assign_cat(name):
    if '와이드' in name:                             return ('리빙', '와이드')
    if '이불 압축' in name or '이불압축' in name:     return ('리빙', '이불')
    if '아우터 압축' in name or '아우터압축' in name:  return ('리빙', '아우터')
    if '행잉 오거나이저' in name or '행잉오거나이저' in name: return ('리빙', '오거나이저')
    if '세이프 데이즈' in name or '데이즈' in name:   return ('여행', '데이즈')
    if '세이프 플러스' in name or '플러스 라인' in name: return ('여행', '플러스')
    if '세이프' in name:                              return ('여행', '세이프')
    if '압축 파우치 라이트' in name:                   return ('여행', '라이트')
    if '여행 압축 파우치' in name:                     return ('여행', '여행')
    if '패커블' in name or '폴더블' in name:           return ('여행', '패커블')
    return ('여행', '기타')

# ── 1. 전체 xlsx 파일 읽기 ──────────────────────────────────────

xlsx_files = sorted(glob.glob(f'{DATA_DIR}/**/*.xlsx', recursive=True) +
                    glob.glob(f'{DATA_DIR}/*.xlsx'))
if not xlsx_files:
    print(f'ERROR: "{DATA_DIR}" 폴더에 .xlsx 파일이 없습니다'); exit(1)

print(f'파일 {len(xlsx_files)}개 발견:')

daily         = defaultdict(lambda: {'s': 0, 'o': 0, 'r': 0, 'v': 0})
prod_all      = defaultdict(lambda: {'s': 0, 'o': 0})
prod_mon      = defaultdict(lambda: defaultdict(lambda: {'s': 0, 'o': 0}))
prod_map      = {}  # code -> {name, s}
prod_code_mon = defaultdict(lambda: defaultdict(int))  # mm -> code -> sales
prod_daily    = defaultdict(lambda: defaultdict(lambda: [0, 0]))  # date -> code -> [sales, orders]

for fname in xlsx_files:
    try:
        wb = openpyxl.load_workbook(fname, data_only=True)
        if 'SALES' not in wb.sheetnames:
            print(f'  SKIP {fname}  (SALES 시트 없음)'); wb.close(); continue
        ws = wb['SALES']
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            if len(row) < 16: continue
            date = to_date(row[0])
            if not date or date == 'None': continue
            mm = date[5:7]
            s, o, r, v = n(row[7]), n(row[3]), n(row[4]), n(row[15])
            d = daily[date]
            d['s'] += s; d['o'] += o; d['r'] += r; d['v'] += v
            name = str(row[1]).strip() if row[1] else ''
            code = str(row[2]).strip() if row[2] else ''
            if name:
                prod_all[name]['s'] += s; prod_all[name]['o'] += o
                prod_mon[mm][name]['s'] += s; prod_mon[mm][name]['o'] += o
            if code and code != 'None' and name:
                if code not in prod_map:
                    prod_map[code] = {'name': name, 's': 0}
                prod_map[code]['s'] += s
                prod_code_mon[mm][code] += int(s)
                prod_daily[date][code][0] += int(s)
                prod_daily[date][code][1] += int(o)
            cnt += 1
        print(f'  OK  {fname}  ({cnt:,}행)')
        wb.close()
    except Exception as e:
        print(f'  ERR {fname}: {e}')

if not daily:
    print('ERROR: 유효한 데이터가 없습니다'); exit(1)

months     = sorted(set(d[5:7] for d in daily.keys()))
start_date = min(daily.keys())
end_date   = max(daily.keys())
print(f'\n기간: {start_date} ~ {end_date}  |  {len(daily)}일  |  월: {", ".join(months)}')

# ── 1b. 방문 분석 xlsx 읽기 ─────────────────────────────────────
MAIN_CHS = ['네이버 서비스', '네이버 광고', '네이버 검색', '직유입']
visit_ch      = defaultdict(lambda: {'v': 0, 'o': 0, 's': 0})
visit_daily   = defaultdict(lambda: defaultdict(lambda: {'v': 0, 'o': 0, 's': 0}))
visit_sub_raw = defaultdict(lambda: defaultdict(lambda: {'v': 0, 'o': 0, 's': 0}))  # date -> "ch1|ch2" -> {v,o,s}
visit_sub_tot = defaultdict(lambda: {'v': 0, 'o': 0, 's': 0})  # "ch1|ch2" -> total
visit_sub3_raw = defaultdict(lambda: defaultdict(lambda: {'v': 0, 'o': 0, 's': 0}))  # date -> "ch1|ch2|ch3" -> {v,o,s}
visit_sub3_tot = defaultdict(lambda: {'v': 0, 'o': 0, 's': 0})  # "ch1|ch2|ch3" -> total

for fname in xlsx_files:
    try:
        wb = openpyxl.load_workbook(fname, data_only=True)
        if 'VISIT' not in wb.sheetnames:
            wb.close(); continue
        ws = wb['VISIT']
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if header is None:
            wb.close(); continue
        # 포맷 감지: 헤더 첫 컬럼이 '날짜 기준'이면 구형, '날짜'면 신형
        h0 = str(header[0]).strip() if header[0] else ''
        old_fmt = '기준' in h0  # 구형: 날짜기준, 날짜, 경로(1단계)... / 신형: 날짜, 경로(1단계)...
        for row in rows_iter:
            if not row or not row[0]: continue
            if old_fmt:
                if str(row[0]).strip() != '조회기간': continue
                date = to_date(row[1])
                ch1  = str(row[2]).strip() if row[2] else ''
                ch2  = str(row[3]).strip() if row[3] else '-'
                ch3  = str(row[4]).strip() if len(row) > 4 and row[4] else '-'
                v_val, o_val, s_val = n(row[5]), n(row[6]), n(row[8])
            else:
                date = to_date(row[0])
                ch1  = str(row[1]).strip() if row[1] else ''
                ch2  = str(row[2]).strip() if row[2] else '-'
                ch3  = str(row[3]).strip() if len(row) > 3 and row[3] else '-'
                v_val, o_val, s_val = n(row[4]), n(row[5]), n(row[7])
            if not date or date == 'None': continue
            if ch1 == '전체': continue
            ch_key = ch1 if ch1 in MAIN_CHS else '기타'
            visit_ch[ch_key]['v'] += v_val
            visit_ch[ch_key]['o'] += o_val
            visit_ch[ch_key]['s'] += s_val
            visit_daily[date][ch_key]['v'] += v_val
            visit_daily[date][ch_key]['o'] += o_val
            visit_daily[date][ch_key]['s'] += s_val
            if ch2 and ch2 != '전체':
                sub_key = f'{ch_key}|{ch2}'
                visit_sub_raw[date][sub_key]['v'] += v_val
                visit_sub_raw[date][sub_key]['o'] += o_val
                visit_sub_raw[date][sub_key]['s'] += s_val
                visit_sub_tot[sub_key]['v'] += v_val
                visit_sub_tot[sub_key]['o'] += o_val
                visit_sub_tot[sub_key]['s'] += s_val
                if ch3 and ch3 != '전체':
                    sub3_key = f'{ch_key}|{ch2}|{ch3}'
                    visit_sub3_raw[date][sub3_key]['v'] += v_val
                    visit_sub3_raw[date][sub3_key]['o'] += o_val
                    visit_sub3_raw[date][sub3_key]['s'] += s_val
                    visit_sub3_tot[sub3_key]['v'] += v_val
                    visit_sub3_tot[sub3_key]['o'] += o_val
                    visit_sub3_tot[sub3_key]['s'] += s_val
        wb.close()
    except Exception:
        pass

# ── 2. JS 블록 생성 ─────────────────────────────────────────────

# ALL: 일별 데이터
rows = []
for d in sorted(daily.keys()):
    lb = f'{d[5:7]}/{d[8:10]}'
    v  = daily[d]
    rows.append(f'  ["{lb}",{int(v["s"])},{int(v["o"])},{int(v["r"])},{int(v["v"])}]')
all_js = 'const ALL = [ // [label, sales, orders, refunds, visits]\n' + ',\n'.join(rows) + '\n];'

# PEAKS: 일평균의 2배 초과 날짜 자동 감지
avg_s = sum(v['s'] for v in daily.values()) / len(daily)
peaks = {f'{d[5:7]}/{d[8:10]}': 1 for d, v in daily.items() if v['s'] > avg_s * 2}
peaks_js = 'const PEAKS = ' + json.dumps(peaks, ensure_ascii=False, separators=(',', ':')) + ';'

# TOP_S / TOP_O
def fmt_list(pairs):
    return '[' + ','.join(f'["{shorten(nm)}",{int(val)}]' for nm, val in pairs) + ']'

def top10(dct, key):
    return [(nm, v[key]) for nm, v in sorted(dct.items(), key=lambda x: x[1][key], reverse=True)[:10]]

s_parts = ['  all:' + fmt_list(top10(prod_all, 's'))]
o_parts = ['  all:' + fmt_list(top10(prod_all, 'o'))]
for mm in months:
    s_parts.append(f"  '{mm}':" + fmt_list(top10(prod_mon[mm], 's')))
    o_parts.append(f"  '{mm}':" + fmt_list(top10(prod_mon[mm], 'o')))
top_s_js = 'const TOP_S = {\n' + ',\n'.join(s_parts) + '\n};'
top_o_js = 'const TOP_O = {\n' + ',\n'.join(o_parts) + '\n};'

# PROD_LIST
pl_rows = []
for code, v in sorted(prod_map.items(), key=lambda x: x[1]['s'], reverse=True):
    nm      = v['name'].replace('\\', '\\\\').replace('"', '\\"')
    maj, mn = assign_cat(v['name'])
    pl_rows.append(f'  ["{code}","{nm}","{maj}","{mn}"]')
prod_list_js = 'const PROD_LIST=[\n' + ',\n'.join(pl_rows) + '\n];'

# PROD_SALES: {all:{code:sales}, '07':{code:sales}, ...}
ps_all = {c: int(v['s']) for c, v in prod_map.items()}
ps_parts = ['  all:' + json.dumps(ps_all, ensure_ascii=False, separators=(',', ':'))]
for mm in months:
    ps_m = {c: int(s) for c, s in prod_code_mon[mm].items()}
    ps_parts.append(f"  '{mm}':" + json.dumps(ps_m, ensure_ascii=False, separators=(',', ':')))
prod_sales_js = 'const PROD_SALES={\n' + ',\n'.join(ps_parts) + '\n};'

# PROD_DAILY: {'MM/DD': {code: [sales, orders]}, ...}
pd_parts = []
for d in sorted(prod_daily.keys()):
    lb = f'{d[5:7]}/{d[8:10]}'
    inner = json.dumps(dict(prod_daily[d]), ensure_ascii=False, separators=(',', ':'))
    pd_parts.append(f'  "{lb}":{inner}')
prod_daily_js = 'const PROD_DAILY={\n' + ',\n'.join(pd_parts) + '\n};'

# VISIT_CH
ch_order = MAIN_CHS + ['기타']
vc_rows = []
for ch in ch_order:
    v = visit_ch.get(ch, {'v': 0, 'o': 0, 's': 0})
    visits = int(v['v'])
    orders = int(v['o'])
    cvr    = round(orders / visits * 100, 1) if visits else 0.0
    sales  = int(v['s'])
    vc_rows.append(f'  ["{ch}",{visits},{orders},{cvr},{sales}]')
visit_ch_js = 'const VISIT_CH=[\n' + ',\n'.join(vc_rows) + '\n];'

# VISIT_DAILY: [label, v0,o0,s0, v1,o1,s1, v2,o2,s2, v3,o3,s3, v4,o4,s4]
vd_rows = []
for d in sorted(visit_daily.keys()):
    lb    = f'{d[5:7]}/{d[8:10]}'
    parts = []
    for ch in ch_order:
        c = visit_daily[d].get(ch, {'v': 0, 'o': 0, 's': 0})
        parts.append(f'{int(c["v"])},{int(c["o"])},{int(c["s"])}')
    vd_rows.append(f'  ["{lb}",{",".join(parts)}]')
visit_daily_js = 'const VISIT_DAILY=[\n' + ',\n'.join(vd_rows) + '\n];'

# VISIT_SUB_DAILY: ch1별 top-N ch2를 날짜별로 기록
# ch1 당 방문 많은 순 top 8 ch2 유지, 나머지는 '기타'로 합산
MAX_SUB = 8
# ch1 별 ch2 랭킹
ch1_top = defaultdict(dict)
for key, tot in visit_sub_tot.items():
    ch1, ch2 = key.split('|', 1)
    ch1_top[ch1][ch2] = tot['v']
ch1_keep = {ch1: set(sorted(d, key=d.get, reverse=True)[:MAX_SUB])
            for ch1, d in ch1_top.items()}

# ch1='직유입', ch2='-' → subchannel 없는 것으로 처리 (skip)
for ch1 in list(ch1_keep.keys()):
    ch2_set = ch1_keep[ch1]
    if ch2_set == {'-'}:
        del ch1_keep[ch1]

# 날짜별 (ch1|ch2) -> {v,o,s}, 기타로 그룹화
grouped = defaultdict(lambda: defaultdict(lambda: {'v': 0, 'o': 0, 's': 0}))
for date, subs in visit_sub_raw.items():
    for sub_key, vals in subs.items():
        ch1, ch2 = sub_key.split('|', 1)
        if ch1 not in ch1_keep:
            continue
        keep_ch2 = ch2 if ch2 in ch1_keep[ch1] else '기타'
        final_key = f'{ch1}|{keep_ch2}'
        grouped[date][final_key]['v'] += vals['v']
        grouped[date][final_key]['o'] += vals['o']
        grouped[date][final_key]['s'] += vals['s']

# 키별 시계열 배열로 변환
sub_series = defaultdict(list)
for date in sorted(grouped.keys()):
    lb = f'{date[5:7]}/{date[8:10]}'
    for final_key, vals in grouped[date].items():
        sub_series[final_key].append([lb, int(vals['v']), int(vals['o']), int(vals['s'])])

vsd_parts = []
for key in sorted(sub_series.keys()):
    inner = json.dumps(sub_series[key], ensure_ascii=False, separators=(',', ':'))
    vsd_parts.append(f'  {json.dumps(key, ensure_ascii=False)}:{inner}')
visit_sub_daily_js = 'const VISIT_SUB_DAILY={\n' + ',\n'.join(vsd_parts) + '\n};'

# VISIT_SUB3_DAILY: ch1|ch2 별 top-N ch3
MAX_SUB3 = 8
ch12_top = defaultdict(dict)
for key, tot in visit_sub3_tot.items():
    parts = key.split('|', 2)
    if len(parts) == 3:
        ch12 = f'{parts[0]}|{parts[1]}'
        ch12_top[ch12][parts[2]] = tot['v']
ch12_keep = {ch12: set(sorted(d, key=d.get, reverse=True)[:MAX_SUB3])
             for ch12, d in ch12_top.items()}
for ch12 in list(ch12_keep.keys()):
    if ch12_keep[ch12] == {'-'}:
        del ch12_keep[ch12]

grouped3 = defaultdict(lambda: defaultdict(lambda: {'v': 0, 'o': 0, 's': 0}))
for date, subs in visit_sub3_raw.items():
    for sub3_key, vals in subs.items():
        parts = sub3_key.split('|', 2)
        if len(parts) != 3: continue
        ch12 = f'{parts[0]}|{parts[1]}'
        ch3 = parts[2]
        if ch12 not in ch12_keep: continue
        keep_ch3 = ch3 if ch3 in ch12_keep[ch12] else '기타'
        final_key = f'{ch12}|{keep_ch3}'
        grouped3[date][final_key]['v'] += vals['v']
        grouped3[date][final_key]['o'] += vals['o']
        grouped3[date][final_key]['s'] += vals['s']

sub3_series = defaultdict(list)
for date in sorted(grouped3.keys()):
    lb = f'{date[5:7]}/{date[8:10]}'
    for final_key, vals in grouped3[date].items():
        sub3_series[final_key].append([lb, int(vals['v']), int(vals['o']), int(vals['s'])])

vsd3_parts = []
for key in sorted(sub3_series.keys()):
    inner = json.dumps(sub3_series[key], ensure_ascii=False, separators=(',', ':'))
    vsd3_parts.append(f'  {json.dumps(key, ensure_ascii=False)}:{inner}')
visit_sub3_daily_js = 'const VISIT_SUB3_DAILY={\n' + ',\n'.join(vsd3_parts) + '\n};'

# ADV_MONTHS: 방문 데이터 월 목록
adv_months = sorted(set(d[5:7] for d in visit_daily.keys()))
adv_month_opts = [f'      <option value="{mm}">{int(mm)}월</option>' for mm in adv_months]
adv_month_opts_html = ('<!-- ADV_MONTH_OPTS -->\n' + '\n'.join(adv_month_opts)
                       + '\n      <!-- /ADV_MONTH_OPTS -->')

# ── 3. dashboard.html 패치 ──────────────────────────────────────

def replace_block(html, start, end, new_js):
    si = html.find(start)
    if si == -1:
        raise ValueError(f'시작 마커를 찾을 수 없음: {repr(start[:50])}')
    ei = html.find(end, si)
    if ei == -1:
        raise ValueError(f'종료 마커를 찾을 수 없음 (시작: {repr(start[:50])})')
    return html[:si] + new_js + html[ei + len(end):]

try:
    with open(DASHBOARD, 'r', encoding='utf-8') as f:
        html = f.read()
except FileNotFoundError:
    print(f'ERROR: {DASHBOARD} 파일이 없습니다. 같은 폴더에서 실행해주세요.'); exit(1)

# 신규 상품 감지: 기존 PROD_LIST 코드 추출
old_codes = set(re.findall(r'\["(\d+)"', html[html.find('const PROD_LIST=['):html.find('\n];', html.find('const PROD_LIST=['))]))
new_codes  = [code for code, _ in sorted(prod_map.items(), key=lambda x: x[1]['s'], reverse=True)]
added      = [c for c in new_codes if c not in old_codes]

# NEW_PRODS JS 생성
new_prods_js = 'const NEW_PRODS=' + json.dumps(added, ensure_ascii=False) + ';'

html = replace_block(html, 'const ALL = [',     '\n];', all_js)
html = replace_block(html, 'const TOP_S = {',   '\n};', top_s_js)
html = replace_block(html, 'const TOP_O = {',   '\n};', top_o_js)
html = replace_block(html, 'const PEAKS = ',    ';',    peaks_js)
html = replace_block(html, 'const NEW_PRODS=',  ';',    new_prods_js)
html = replace_block(html, 'const PROD_SALES={\n', '\n};', prod_sales_js)
html = replace_block(html, 'const PROD_DAILY={\n', '\n};', prod_daily_js)
html = replace_block(html, 'const PROD_LIST=[\n', '\n];', prod_list_js)
html = replace_block(html, 'const VISIT_CH=[\n', '\n];', visit_ch_js)
html = replace_block(html, 'const VISIT_DAILY=[\n', '\n];', visit_daily_js)
html = replace_block(html, 'const VISIT_SUB_DAILY={\n', '\n};', visit_sub_daily_js)
html = replace_block(html, 'const VISIT_SUB3_DAILY={\n', '\n};', visit_sub3_daily_js)
if adv_months:
    html = replace_block(html, '<!-- ADV_MONTH_OPTS -->', '<!-- /ADV_MONTH_OPTS -->', adv_month_opts_html)
    adv_start = min(visit_daily.keys())
    adv_end   = max(visit_daily.keys())
    adv_last30 = max(adv_start, (datetime.strptime(adv_end, '%Y-%m-%d') - timedelta(days=29)).strftime('%Y-%m-%d'))
    html = re.sub(r'(<input type="date" id="advFrom")[^>]*(>)',
                  f'\\1 value="{adv_last30}" min="{adv_start}" max="{adv_end}"\\2', html)
    html = re.sub(r'(<input type="date" id="advTo")[^>]*(>)',
                  f'\\1 value="{adv_end}" min="{adv_start}" max="{adv_end}"\\2', html)

# 월별 필터 select 옵션 갱신
month_opts = []
for mm in months:
    month_opts.append(f'      <option value="{mm}">{int(mm)}월</option>')
month_opts_html = '<!-- MONTH_OPTS -->\n' + '\n'.join(month_opts) + '\n      <!-- /MONTH_OPTS -->'
html = replace_block(html, '<!-- MONTH_OPTS -->', '<!-- /MONTH_OPTS -->', month_opts_html)

# 날짜 입력 필드 min/max/value 업데이트
yr = start_date[:4]
sorted_months = sorted(months)
# 비교 탭 기본값: 마지막 두 달
if len(sorted_months) >= 2:
    last_m, prev_m = sorted_months[-1], sorted_months[-2]
    days_prev = calendar.monthrange(int(yr), int(prev_m))[1]
    p1_from = f'{yr}-{prev_m}-01'
    p1_to   = f'{yr}-{prev_m}-{days_prev:02d}'
    p2_from = f'{yr}-{last_m}-01'
    p2_to   = end_date
else:
    p1_from = p1_to = p2_from = p2_to = start_date

def upd_input(html, id_, value, min_, max_):
    return re.sub(
        rf'(<input type="date" id="{id_}")[^>]*(>)',
        f'\\1 value="{value}" min="{min_}" max="{max_}"\\2',
        html
    )

a_last30 = max(start_date, (datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=29)).strftime('%Y-%m-%d'))
html = upd_input(html, 'aFrom',  a_last30,   start_date, end_date)
html = upd_input(html, 'aTo',    end_date,   start_date, end_date)
html = upd_input(html, 'c1From', p1_from,    start_date, end_date)
html = upd_input(html, 'c1To',   p1_to,      start_date, end_date)
html = upd_input(html, 'c2From', p2_from,    start_date, end_date)
html = upd_input(html, 'c2To',   p2_to,      start_date, end_date)

# 타이틀 날짜 범위 업데이트 (예: 2026.07–08)
m1, m2 = start_date[5:7], end_date[5:7]
new_range = f'{yr}.{m1}–{m2}' if m1 != m2 else f'{yr}.{m1}'
html = re.sub(r'(매출 추이 대시보드 · )[\d.–\-]+', r'\g<1>' + new_range, html)

# PROD_LIST 주석 개수 업데이트
html = re.sub(
    r'// PRODUCT LIST DATA \(\d+개 상품\)',
    f'// PRODUCT LIST DATA ({len(pl_rows)}개 상품)',
    html
)

with open(DASHBOARD, 'w', encoding='utf-8') as f:
    f.write(html)
with open(INDEX, 'w', encoding='utf-8') as f:
    f.write(html)

print(f'\n✓ {DASHBOARD} + {INDEX} 업데이트 완료')
print(f'  일별: {len(rows)}일  |  피크: {len(peaks)}일  |  상품: {len(pl_rows)}개')
if added:
    print(f'\n  ★ 신규 상품 {len(added)}개 — 대시보드에서 카테고리 지정 필요:')
    for c in added:
        print(f'    - [{c}] {prod_map[c]["name"]}')
else:
    print(f'  신규 상품 없음')
